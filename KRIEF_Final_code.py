#%%
import csv
from itertools import zip_longest
import numpy as np
from math import *
import matplotlib.pyplot as plt
import openpyxl as xl
from openpyxl import Workbook
import random
S1=[]
S2=[]
S3=[]
S4=[]
S5=[]
timeserie=[]
with open('/Users/bryankrief/Desktop/Cours/Skema/FMI - S1/Python/final python s2/FINAL.csv') as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    line_count = 0
    for row in csv_reader:
        if line_count == 0:
            line_count += 1
        else:
            if line_count == 1:
                line_count += 1
            else:
                #retrieve stocks and dates to lists
                line_count += 1
                S1.append(float(row[1]))
                S2.append(float(row[2]))
                S3.append(float(row[3]))
                S4.append(float(row[4]))
                S5.append(float(row[5]))
                if line_count>1:
                    timeserie.append(row[0])
#function to get the daily log return of a list
def returnX(serie):
    compte=0
    ret=[]
    for i in serie:
        if compte==0:
            ret1=i
        else:
            return1=log(i/ret1)
            ret.append(return1)
            ret1=i
        compte+=1
    return ret
#assigning functions to new list
returnlist1=returnX(S1)
returnlist2=returnX(S2)
returnlist3=returnX(S3)
returnlist4=returnX(S4)
returnlist5=returnX(S5)

#merging lists
d = [timeserie,returnlist1, returnlist2,returnlist3,returnlist4,returnlist5]
export_data = zip_longest(*d, fillvalue = '')
#write in characteristics.csv the merged lists
with open('Characteristics.csv','w',newline='') as f:
    writein=csv.writer(f)
    writein.writerow(['Time','Stock 1','Stock 2','Stock 3','Stock 4','Stock 5'])
    for y in export_data:
        writein.writerow(y)

#function to get random allocation of portfolio 
def weighting(number_stocks):
    a=[]
    for i in range(number_stocks):
        a.append(np.random.uniform())
    return a

ptf_mean=[]
ptf_std=[]

for z in range(10000):
    #get new allocations 
    weight_matrix=[]
    weight_matrix=weighting(5)
    matrix_total=np.sum(weight_matrix)
    #allocations converted in percentage
    stok1_w=weight_matrix[0]/matrix_total
    stok2_w=weight_matrix[1]/matrix_total
    stok3_w=weight_matrix[2]/matrix_total
    stok4_w=weight_matrix[3]/matrix_total
    stok5_w=weight_matrix[4]/matrix_total
    #assign results to new variables sum(allocation weighting*mean/std of the lists)
    portfolio_result_mean=stok1_w*np.mean(returnlist1)+stok2_w*np.mean(returnlist2)+stok3_w*np.mean(returnlist3)+stok4_w*np.mean(returnlist4)+stok5_w*np.mean(returnlist5)
    portfolio_result_std=stok1_w*np.std(returnlist1)+stok2_w*np.std(returnlist2)+stok3_w*np.std(returnlist3)+stok4_w*np.std(returnlist4)+stok5_w*np.std(returnlist5)
    ptf_mean.append(portfolio_result_mean*100)
    ptf_std.append(portfolio_result_std*100)
#print results in xl workbook saved as .csv
wb= Workbook()
ws1=wb.active
ws1.cell(1,1).value = "Mean of daily returns"
ws1.cell(1,2).value = "Mean of daily volatility"
ws1.cell(1,3).value = "(results in %)"
ws1.cell(1,4).value = "Mean of daily returns unrestricted"
ws1.cell(1,5).value = "Mean of daily volatility unrestricted"
ws1.cell(1,6).value = "(results in %)"
r=2
for affich in ptf_mean:
    ws1.cell(r,1).value = affich
    r=r+1
r=2
for affich in ptf_std:
    ws1.cell(r,2).value = affich
    r=r+1

#plot results
plt.scatter( ptf_std, ptf_mean, color='r')
plt.xlabel('Daily volatility (in %)')
plt.ylabel('Mean of daily returns (in %)')
plt.show()

#repeat the operation with no restrictions------------
#function to get random allocation of portfolio no restriction s
def weighting_unrestricted(number_stocks):
    a=[]
    for i in range(number_stocks):
        a.append(random.uniform(-1,1))
    return a

ptf_mean_nr=[]
ptf_std_nr=[]
for z in range(10000):
    #get new allocations 
    weight_matrix=[]
    weight_matrix=weighting_unrestricted(5)
    matrix_total=np.sum(weight_matrix)
    #allocations converted in percentage
    stok1_w=weight_matrix[0]/matrix_total
    stok2_w=weight_matrix[1]/matrix_total
    stok3_w=weight_matrix[2]/matrix_total
    stok4_w=weight_matrix[3]/matrix_total
    stok5_w=weight_matrix[4]/matrix_total
    #assign results to new variables sum(allocation weighting*mean/std of the lists)
    portfolio_result_mean=stok1_w*np.mean(returnlist1)+stok2_w*np.mean(returnlist2)+stok3_w*np.mean(returnlist3)+stok4_w*np.mean(returnlist4)+stok5_w*np.mean(returnlist5)
    portfolio_result_std=stok1_w*np.std(returnlist1)+stok2_w*np.std(returnlist2)+stok3_w*np.std(returnlist3)+stok4_w*np.std(returnlist4)+stok5_w*np.std(returnlist5)
    ptf_mean_nr.append(portfolio_result_mean*100)
    ptf_std_nr.append(portfolio_result_std*100)
r=2
for affich in ptf_mean:
    ws1.cell(r,4).value = affich
    r=r+1
r=2
for affich in ptf_std:
    ws1.cell(r,5).value = affich
    r=r+1
wb.save(filename="/Users/bryankrief/Desktop/Cours/Skema/FMI - S1/Python/final python s2/Whatifportfolios.csv")
#plot results
plt.scatter( ptf_std_nr, ptf_mean_nr, color='b')
plt.xlabel('Daily volatility (in %)')
plt.ylabel('Mean of daily returns (in %)')
plt.show()
#plot all results
plt.scatter( ptf_std_nr, ptf_mean_nr, color='b')
plt.scatter( ptf_std, ptf_mean, color='r')
plt.xlabel('Daily volatility (in %)')
plt.ylabel('Mean of daily returns (in %)')
plt.show()
#Note: results of whatif portfolios are in xl and saved as .csv (because requested) which means that you have still to open it under excel to see it properly
#%%
